"""
excel_export.py

Small shared helper for turning a fee table into a downloadable, nicely
formatted .xlsx workbook -- used by the Fee Estimate tab's two "Export to
Excel" buttons (the hours x rate discipline build-up, and the AI benchmark
% split table). Kept separate from the fee-domain modules (resourcing.py,
fee_estimation_engine.py) since this is presentation, not fee logic -- the
same reasoning divider_designer.py sits apart from export_docx.py.

openpyxl is an optional dependency here on purpose: if it isn't installed,
build_fee_workbook() returns None rather than raising, so a missing package
degrades to "the export button shows an install hint" instead of crashing
the whole app.
"""

from __future__ import annotations

import io

from modules.divider_designer import THEME_COLOURS


def build_fee_workbook(
    sheet_title: str,
    headers: list[str],
    rows: list[list],
    column_formats: dict[int, str] | None = None,
    summary_rows: list[list] | None = None,
    theme_name: str | None = None,
    title: str | None = None,
    meta: list[tuple[str, str]] | None = None,
    notes: list[str] | None = None,
) -> bytes | None:
    """
    Build a single-sheet .xlsx workbook: a themed bold header row (using the
    same THEME_COLOURS every generated graphic in this tool uses, so the
    export at least nods to the proposal's colours), the data rows with
    per-column number formats, optional summary rows (e.g. totals, an
    average-rate figure) rendered bold with a light fill so they read as
    distinct from the data rather than just another row, and auto-sized
    columns. Returns .xlsx bytes, or None if openpyxl isn't installed.

    title/meta render an identifying block above the table -- what project
    this fee is for, which client, the tender number, when it was exported.
    Without it these workbooks left the app as an anonymous grid of numbers:
    once one is emailed to a colleague or saved next to three other bids,
    nothing in the file says which tender it belongs to.

    notes render below the table in italics -- used to say what a blank cell
    means, which matters when blank is deliberate.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    try:
        colours = THEME_COLOURS.get(theme_name, THEME_COLOURS["Corporate"])
        primary = colours["primary"]
        header_rgb = "%02X%02X%02X" % primary
        light_bg = sum(primary) / 3 > 150
        header_font_colour = "222222" if light_bg else "FFFFFF"

        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_title or "Sheet1")[:31]  # Excel caps sheet names at 31 chars

        header_font = Font(bold=True, color=header_font_colour, size=11)
        header_fill = PatternFill("solid", fgColor=header_rgb)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        r = 1
        if title:
            cell = ws.cell(row=r, column=1, value=title)
            cell.font = Font(bold=True, size=14, color=header_rgb)
            r += 1
        for label, value in (meta or []):
            ws.cell(row=r, column=1, value=f"{label}:").font = Font(bold=True)
            ws.cell(row=r, column=2, value=value)
            r += 1
        if title or meta:
            r += 1  # one blank row between the identifying block and the table

        header_row = r
        for col, column_title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.freeze_panes = f"A{header_row + 1}"

        r = header_row + 1
        for row in rows:
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.border = border
                fmt = (column_formats or {}).get(col)
                if fmt and value is not None:
                    cell.number_format = fmt
            r += 1

        all_rows = list(rows)
        if summary_rows:
            summary_font = Font(bold=True)
            summary_fill = PatternFill("solid", fgColor="F2F2F2")
            for row in summary_rows:
                for col, value in enumerate(row, start=1):
                    cell = ws.cell(row=r, column=col, value=value)
                    cell.font = summary_font
                    cell.fill = summary_fill
                    cell.border = border
                    fmt = (column_formats or {}).get(col)
                    if fmt and value is not None:
                        cell.number_format = fmt
                r += 1
            all_rows = all_rows + summary_rows

        if notes:
            r += 1
            note_font = Font(italic=True, size=9, color="808080")
            for note in notes:
                ws.cell(row=r, column=1, value=note).font = note_font
                r += 1

        # openpyxl has no built-in autofit -- estimate column width from the
        # longest string (header or content) in that column.
        for col in range(1, len(headers) + 1):
            longest = len(str(headers[col - 1]))
            if col == 1:
                # The identifying block's labels live in column 1 too, and a
                # note is much longer than any fee row -- size to the labels,
                # not to the notes, or column A ends up 60 characters wide.
                for label, _value in (meta or []):
                    longest = max(longest, len(f"{label}:"))
            for row in all_rows:
                if col - 1 < len(row) and row[col - 1] is not None:
                    longest = max(longest, len(str(row[col - 1])))
            ws.column_dimensions[get_column_letter(col)].width = min(max(longest + 2, 10), 60)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    except Exception:
        return None
